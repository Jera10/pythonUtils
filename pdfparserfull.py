import sys
import re
import json
import openpyxl

def parse_skus_and_amounts(text):
    lines = text.splitlines()
    results = []
    current_sku = None

    i = 0
    while i < len(lines):
        line = lines[i]

        # Stop processing if "Removed" is found
        if "Removed" in line:
            break

        # Check for SKU
        sku_match = re.search(r'SKU:\s*([\w-]+)', line)
        if sku_match:
            current_sku = sku_match.group(1)

        # Check for amounts in this line
        amount_matches = re.findall(r'×\s*(\d+)', line)
        if amount_matches:
            # Determine which SKU to use
            if current_sku:
                sku_for_this = current_sku
                current_sku = None  # Consume the SKU after using
            else:
                # Look ahead for SKU if none seen yet
                lookahead_sku = None
                for j in range(1, 4):  # Look ahead up to 3 lines
                    if i + j < len(lines):
                        future_line = lines[i + j]
                        future_match = re.search(r'SKU:\s*([\w-]+)', future_line)
                        if future_match:
                            lookahead_sku = future_match.group(1)
                            break
                sku_for_this = lookahead_sku if lookahead_sku else "Missing article"

            for amt in amount_matches:
                results.append({
                    "SKU": sku_for_this,
                    "Amount": int(amt)
                })
        i += 1

    # Post-processing: fix consecutive duplicate SKUs
    for i in range(1, len(results)):
        if results[i]["SKU"] == results[i - 1]["SKU"]:
            results[i - 1]["SKU"] = "Missing article"

    return results

# Get file path from command-line argument
file_path = sys.argv[1]

# Try reading with UTF-16, fallback to UTF-8
try:
    with open(file_path, 'r', encoding='utf-16') as f:
        text_content = f.read()
except UnicodeDecodeError:
    with open(file_path, 'r', encoding='utf-8') as f:
        text_content = f.read()

# Parse the data
parsed_data = parse_skus_and_amounts(text_content)

# Print JSON to console (optional)
print(json.dumps(parsed_data, indent=2))

# Write to Excel

excel_path = sys.argv[2]

wb = openpyxl.load_workbook(excel_path, keep_vba=True)
sheet = wb["specification merchandise"]

start_row = 4
for i, item in enumerate(parsed_data):
    sheet.cell(row=start_row + i, column=2, value=item["SKU"])    # Column B
    sheet.cell(row=start_row + i, column=6, value=item["Amount"]) # Column F

wb.save(excel_path)
